import openpyxl
import re
import os
import sys
from datetime import datetime

VALID_INPUT = ["1", "2", "3", "4", "5", 1, 2, 3, 4, 5]

if "--debug" in sys.argv:
    DEBUG = True
else:
    DEBUG = False

def errorRouter():
    input("Pressione ENTER para sair")
    sys.exit(0)

class ESPM_Student_ST:
    def __init__(self, RA, info=None):
        self.RA = RA
        self.traces = {}
        if info is not None:
            self.nome = info["Nome"]
            self.turma = info["Turma"]
            self.semestre = info["Semestre"]
        else:
            self.nome = None
            self.turma = None
            self.semestre = None

    def add(self, trace_name, trace_value):
        if trace_name not in self.traces:
            self.traces[trace_name] = trace_value
            print("\t\t%s : %s" %(trace_name, trace_value))
        else:
            print("\t\tTentativa de redefinição de nota (%s). Ignorando." %(trace_name))

class ESPM_Student_WS:
    def __init__(self, ws, filename):
        self.ws = ws
        self.structure = self.detectStructure()
        self.filename = filename
        self.jsonDatafile = None

    def readStudents(self):
        data = {}
        if self.structure["Header"]["row"] is not None:
            for row in range(self.structure["Header"]["row"] + 1, self.ws.max_row + 1):
                RA = self.ws[ self.structure["RA"]["column"] + str(row) ].value
                if RA is not None:
                    RA = str(RA)
                    data[RA] = {"info":
                                    {   "Nome":None,
                                        "Turma":None,
                                        "Semestre":None
                                    },
                                "tracos": {}
                                }
                    if self.structure["Nome"]["column"] is not None:
                        data[RA]["info"]["Nome"] = self.ws[ self.structure["Nome"]["column"] + str(row) ].value
                    if self.structure["Turma"]["column"] is not None:
                        data[RA]["info"]["Turma"] = self.ws[ self.structure["Turma"]["column"] + str(row) ].value
                    if self.structure["Semestre"]["column"] is not None:
                        data[RA]["info"]["Semestre"] = self.ws[ self.structure["Semestre"]["column"] + str(row) ].value


                    for traco in self.structure["Traços"]:
                        traco_value = self.ws[ self.structure["Traços"][traco] + str(row) ].value
                        if traco_value in VALID_INPUT:
                            data[RA]["tracos"][traco] = traco_value

        return data


    def detectStructure(self):
        ws_structure = {"RA":{"column":None},
                        "Nome":{"column":None},
                        "Turma":{"column":None},
                        "Semestre":{"column":None},
                        "Traços":{},
                        "Header":{"row":None},
                        "Professor":None}
        for row in self.ws.iter_rows():
            nextIsProfessorName = False
            for cell in row:
                if type(cell.value) is str:
                    if "Professor" == cell.value:
                        nextIsProfessorName = True
                    elif nextIsProfessorName:
                        nextIsProfessorName = False
                        ws_structure["Professor"] = cell.value
                    elif "R.A" == cell.value or "RA" == cell.value:
                        ws_structure["RA"]["column"] = cell.column
                        ws_structure["Header"]["row"] = cell.row
                    elif "Nome" == cell.value:
                        ws_structure["Nome"]["column"] = cell.column
                    elif "Turma" == cell.value:
                        ws_structure["Turma"]["column"] = cell.column
                    elif "Semestre" == cell.value:
                        ws_structure["Semestre"]["column"] = cell.column
                    elif re.match('C(([0-9]{1,2}(()|( ))T[0-9]{1})|[A-Z]{3})', cell.value):
                        ws_structure["Traços"][cell.value] = cell.column
        return ws_structure

def main():
    print("#####################################################################")
    print("#                      LEITOR DE PLANILHAS ESPM                     #")
    print("#                                                                   #")
    print("# 1) Certifique-se que todas as planilhas estão no diretório /aval  #")
    print("# 2) Certifique-se que todas as planilhas têm um formato válido     #")
    print("# 3) Certifique-se o arquivo tracos.txt lista todos os tracos       #")
    print("#    avaliados.                                                     #")
    print("# 4) O resultado da coleta será gravado no arquivo output.xlsx      #")
    print("#    Certifique-se que ele não está aberto em outro programa        #")
    print("#                                                                   #")
    print("# Suporte: Renan Yuri Lino (renan.lino@gmail.com)                   #")
    print("#####################################################################")
    input("Pressione ENTER para começar...")

    os.chdir("..")
    print(os.getcwd())

    print("Lendo traços...")
    try:
        listaTracos = open("tracos.txt", "r")
    except IOError:
        print("400 - Erro ao abrir a referência de competências e traços em tracos.txt. Abortando.")
        errorRouter()

    hdr = ["N", "Avaliador", "RA", "Amostra/Estudantes", "Turma",
            "Semestre do estudante", "Data da coleta"]
    vetorTracos = []
    for line in listaTracos:
        line = line.replace("\n","").replace("\r","").replace(" ", "")
        hdr.append(line)
        vetorTracos.append(line)
    listaTracos.close()

    print()
    for t in vetorTracos:
        print("\t" + t)
    print("Leu %d traços" %len(vetorTracos))
    input("Pressione ENTER para continuar...")

    outputWb = openpyxl.Workbook(write_only=True)
    wsOut = outputWb.create_sheet()

    wsOut.append(hdr)

    timestamp = datetime.now()
    semester = timestamp.month // 6 + 1
    nowStamp = timestamp.strftime("%Y")
    nowStamp += "." + str(semester)

    os.chdir("./aval")
    print()
    print("Lendo planilhas de %s" %os.getcwd())
    for filename in os.listdir():
        if re.match('.*\.xlsx', filename) and "~$" not in filename:
            print("\t" + filename)
    print("Encontrou %d planilhas." %len(os.listdir()))
    input("Pressione ENTER para continuar...")
    N = 1
    for filename in os.listdir():
        if re.match('.*\.xlsx', filename) and "~$" not in filename:
            try:
                wb = openpyxl.load_workbook(filename)
            except IOError:
                print("401 - Erro ao abrir o arquivo %s. Ignorando." %(filename))
                continue
            for sheet_name in wb.get_sheet_names():
                if "Traços" not in sheet_name:

                    ws_data = ESPM_Student_WS(wb[sheet_name], filename)
                    avaliador = ws_data.structure["Professor"].title()
                    print("Processando avaliações de %s (%s)" %(avaliador, sheet_name ))
                    data = ws_data.readStudents()

                    for RA in data:
                        entry = [str(N), str(avaliador), str(RA),
                                str(data[RA]["info"]["Nome"]),
                                str(data[RA]["info"]["Turma"]),
                                str(data[RA]["info"]["Semestre"]),
                                nowStamp]
                        offset = len(entry)
                        for i in range(len(hdr) - len(entry)):
                            entry.append("")
                        if DEBUG:
                            print("\t%s (%s) | T: %s | Sem: %s" %(entry[3], entry[2],
                                        entry[4], entry[5]) )
                        for traco in data[RA]["tracos"]:
                            if DEBUG:
                                print("\t\t%s : %s" %(traco, data[RA]["tracos"][traco]))
                            try:
                                i = hdr.index(traco)
                            except ValueError:
                                print("\t\t403 - Competência/traço %s inesperado. Ignorando." %traco)
                                continue
                            entry[i] = str(data[RA]["tracos"][traco])

                        wsOut.append(entry)
                        N += 1

    os.chdir("../")
    try:
        outputWb.save('output.xlsx')
    except IOError:
        print("405 - Erro ao salvar o arquivo output.xlsx. Ele está aberto em outro programa?")
        errorRouter()

    print("Coleta salva em output.xlsx. Encerrando.")

main()
input()
